# -*- coding: utf-8 -*-
# Синхронизировано по функционалу с исходными процедурами из output_excel_old.py:
# create_title_sheet, rider-блок, Phito_Check, Finish_Check и утилиты
# См. источники: :contentReference[oaicite:0]{index=0} :contentReference[oaicite:1]{index=1} :contentReference[oaicite:2]{index=2}

import os
import math
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

class ExcelExporter:
    """
    Класс экспортирует все листы Excel и полностью повторяет функционал исходного модуля (old):
      - Conos_Check (rider-таблица + суммы + доп.значения + подсветки/проверки)
      - Phito_Check (свод проверок фитосанитарки)
      - Finish_Check (сравнение коносамента и фитосертификата, контейнеры, подсветки)
      - Conos_Data (title/rider в развернутом виде)
      - Phytosanitary Data (страницы 1/2)
    Параметры конструктора соответствуют глобальным переменным исходника.
    """

    def __init__(self, data_processor, EXCEL_PATH):
        self.data_processor = data_processor
        self.EXCEL_PATH = EXCEL_PATH
        
        # Ссылки на все данные из data_processor
        self.extract_conos = data_processor.extract_conos
        self.extract_phito = data_processor.extract_phito
        self.data_conos = data_processor.data_conos
        self.data_phito = data_processor.data_phito
        self.amounts_conos = data_processor.amounts_conos
        self.rows_count = data_processor.rows_count
        self.check_in_conos = data_processor.check_in_conos
        self.check_in_phito = data_processor.check_in_phito
        self.finish_conos = data_processor.finish_conos
        self.finish_phito = data_processor.finish_phito
        self.control_results = data_processor.control_results

        self.styles = self.create_styles()

    # ------------------------- Публичный API -------------------------

    def export_to_excel(self):
        """Основная точка входа: формирует и сохраняет Excel-файл."""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1) Conos_Check (rider-данные)
        ws_conos_check = wb.create_sheet("Conos_Check", 0)

        max_len = len(self.extract_conos["containers"])
        boxes = self.align_list(self.extract_conos["boxes"], max_len, 0)
        per_box = self.align_list(self.extract_conos["perbox"], max_len, 0.0)
        weights = self.align_list(
            [float(w) if w else 0.0 for w in self.extract_conos["weights"]],
            max_len,
            0.0,
        )
        fresh = self.align_list(self.extract_conos["fresh"], max_len, "N/A")
        fruits = self.align_list(self.extract_conos["fruits"], max_len, "N/A")

        self.add_rider_headers(ws_conos_check)
        self.add_row_counters(ws_conos_check)
        self.fill_rider_data(ws_conos_check, boxes, per_box, weights, fresh, fruits)
        self.add_rider_totals(ws_conos_check)
        self.add_rider_additional_values(ws_conos_check)
        self.apply_rider_checks(ws_conos_check)
        self.auto_adjust_columns(ws_conos_check)

        # 2) Phito_Check
        if self.data_phito:
            self.create_system_info_sheet(wb)

        # 3) Finish_Check
        self.create_finish_check_sheet(wb)

        # 4) Conos_Data
        self.create_title_sheet(wb)

        # 5) Phytosanitary Data
        if self.data_phito:
            self.create_phytosanitary_sheet(wb)

        wb.save(self.EXCEL_PATH)
        try:
            os.startfile(self.EXCEL_PATH)  # Windows
        except Exception:
            # На *nix системах просто игнорируем
            pass

    # ------------------------- Стили -------------------------

    def create_styles(self):
        return {
            "green_fill": PatternFill(
                start_color="00C800", end_color="00C800", fill_type="solid"
            ),
            "orange_fill": PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            ),
            "header_fill": PatternFill(
                start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
            ),
            "bold_font": Font(color="000000", bold=True),
            "wrap_alignment": Alignment(wrap_text=True, vertical="top"),
            "center_alignment": Alignment(horizontal="center", vertical="center"),
        }

    # ------------------------- Conos_Data -------------------------

    def create_title_sheet(self, wb):
        """Conos_Data: вывод полей Title и Rider из коносамента с форматированием."""
        if "Conos_Data" in wb.sheetnames:
            ws_title = wb["Conos_Data"]
            ws_title.delete_rows(1, ws_title.max_row)
        else:
            ws_title = wb.create_sheet("Conos_Data")

        header_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws_title.append(["TITLE DATA"])
        ws_title.merge_cells("A1:B1")
        title_cell = ws_title["A1"]
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="center")

        # Title
        for (page_num, block_num), block_data in sorted(self.data_conos["title"].items()):
            ws_title.append(
                [f"Block {block_num}: {block_data['name']}", block_data["value"]]
            )

        # Rider
        ws_title.append([])
        ws_title.append(["RIDER DATA"])
        ws_title.merge_cells(f"A{ws_title.max_row}:B{ws_title.max_row}")
        rider_cell = ws_title[f"A{ws_title.max_row}"]
        rider_cell.font = Font(bold=True, size=14)
        rider_cell.fill = header_fill
        rider_cell.alignment = Alignment(horizontal="center")

        for (page_num, block_num), block_data in sorted(self.data_conos["rider"].items()):
            ws_title.append(
                [f"Page {page_num} Block {block_num}: {block_data['name']}", block_data["value"]]
            )

        # Размеры и высота строк
        ws_title.column_dimensions["A"].width = 45
        ws_title.column_dimensions["B"].width = 120

        for row_idx, row in enumerate(ws_title.iter_rows(), 1):
            if row_idx == 1:
                ws_title.row_dimensions[row_idx].height = 25
                continue
            if not any(cell.value for cell in row):
                ws_title.row_dimensions[row_idx].height = 15
                continue

            max_lines = 1
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    line_count = len(lines)
                    wrapped_lines = sum(math.ceil(len(line) / 100) for line in lines)
                    max_lines = max(max_lines, line_count, wrapped_lines)

            row_height = min(15 + (max_lines * 12), 400)
            ws_title.row_dimensions[row_idx].height = row_height

            for cell in row:
                cell.alignment = wrap_alignment
                cell.border = border

        if ws_title.max_row > 1:
            ws_title.auto_filter.ref = f"A2:B{ws_title.max_row}"

        return ws_title

    # ------------------------- Conos_Check (Rider) -------------------------

    def add_rider_headers(self, ws):
        headers = [
            "Контейнеры",
            "BOX(ES)",
            "PER BOX (KGS)",
            "Вес (KGS)",
            "Total Weight (KGS)",
            "Fresh Count",
            "Fruits",
        ]
        ws.append(headers)
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            cell.font = self.styles["bold_font"]
            if col == 5:  # Total Weight
                cell.fill = self.styles["header_fill"]

    def add_row_counters(self, ws):
        ws.append(
            [
                self.rows_count["conos"]["total_cont"],
                self.rows_count["conos"]["total_boxes"],
                self.rows_count["conos"]["total_perbox"],
                self.rows_count["conos"]["total_weights"],
                "",  # Total Weight (KGS)
                self.rows_count["conos"]["total_fresh"],
                self.rows_count["conos"]["total_fruits"],
            ]
        )

    def fill_rider_data(self, ws, boxes, per_box, weights, fresh, fruits):
        for i in range(len(self.extract_conos["containers"])):
            total_weight = (
                boxes[i] * per_box[i] if (i < len(boxes)) and boxes[i] != 0 else 0.0
            )

            row_data = [
                self.extract_conos["containers"][i]
                if i < len(self.extract_conos["containers"])
                else "",
                boxes[i],
                per_box[i],
                weights[i],
                float(total_weight),
                fresh[i],
                fruits[i],
            ]
            ws.append(row_data)

            current_row = ws.max_row
            # Подсветка Total Weight
            ws.cell(row=current_row, column=5).fill = self.styles["header_fill"]
            # Нулевые значения
            if boxes[i] == 0:
                ws.cell(row=current_row, column=2).fill = self.styles["orange_fill"]
            if per_box[i] == 0.0 and i >= self.rows_count["conos"]["total_perbox"]:
                ws.cell(row=current_row, column=3).fill = self.styles["orange_fill"]

    def add_rider_totals(self, ws):
        last_row = ws.max_row
        formula_row = last_row + 1
        formulas = [
            f"=COUNTA(A3:A{last_row})",
            self.amounts_conos["SUM_boxes"],
            self.amounts_conos["MID_perbox"],
            self.amounts_conos["SUM_weights"],
            self.amounts_conos["AM_total_weight"],  # итог по Total Weight
            f"=COUNTA(F3:F{last_row})",
            f"=COUNTA(G3:G{last_row})",
        ]
        for col, formula in enumerate(formulas, 1):
            cell = ws.cell(row=formula_row, column=col, value=formula)
            cell.fill = self.styles["header_fill"]
            cell.font = self.styles["bold_font"]

    def add_rider_additional_values(self, ws):
        """Доп. значения EX9/EX13/EX14 + форматирование весовых столбцов."""
        formula_row = ws.max_row
        additional_row_1 = formula_row + 1
        additional_row_2 = formula_row + 2

        ex9_value = self.safe_int(self.extract_conos["EX9_qty_CONT_CARRIERS"])
        ex13_value = self.safe_int(self.extract_conos["EX13_qty_CONT"])
        ex14_value = self.safe_float(self.extract_conos["EX14_total_GCW"])
        ex13_total_items = self.safe_int(self.extract_conos.get("EX13_total_items", "0"))

        ws.cell(row=additional_row_1, column=1, value=ex9_value).comment = Comment(
            f"EX9_qty_CONT_CARRIERS = {ex9_value}", "Система"
        )
        ws.cell(row=additional_row_2, column=1, value=ex13_value).comment = Comment(
            f"EX13_qty_CONT = {ex13_value}", "Система"
        )

        cell_ex14 = ws.cell(row=additional_row_1, column=4, value=ex14_value)
        cell_ex14.comment = Comment(f"EX14_total_GCW = {ex14_value}", "Система")
        cell_ex14.number_format = "0.00"

        ws.cell(row=additional_row_1, column=2, value=ex13_total_items).comment = Comment(
            f"EX13_total_items = {ex13_total_items}", "Система"
        )

        # Формат чисел для C/D/E
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = "0.00"

    def apply_rider_checks(self, ws):
        """Подсветка контрольных значений и результатов проверок (Conos_Check)."""
        self.run_check_control()

        # строка с формулами — + две добавленные строки
        formula_row = ws.max_row - 2

        # Контрольные ячейки под итогами
        controls = [
            (formula_row + 1, 1, "EX9_is_number"),
            (formula_row + 2, 1, "EX13_qty_CONT"),
            (formula_row + 1, 2, "EX13_total_items"),
            (formula_row + 1, 4, "EX14_total_GCW"),
        ]
        for row, col, key in controls:
            cell = ws.cell(row=row, column=col)
            cell.fill = (
                self.styles["green_fill"]
                if self.check_in_conos.get(key)
                else self.styles["orange_fill"]
            )

        # Поэлементные проверки веса
        for i, is_valid in self.check_in_conos.get("weights", {}).items():
            cell = ws.cell(row=3 + i, column=4)
            cell.fill = self.styles["green_fill"] if is_valid else self.styles["orange_fill"]

        # Сверка количества строк (ROWS)
        rows_mapping = {
            "total_cont": 1,
            "total_boxes": 2,
            "total_perbox": 3,
            "total_weights": 4,
            "total_fresh": 6,
            "total_fruits": 7,
        }

        if self.check_in_conos.get("rows_equally", False):
            for col in rows_mapping.values():
                ws.cell(row=2, column=col).fill = self.styles["green_fill"]
        elif "rows_individual" in self.check_in_conos:
            for key, col in rows_mapping.items():
                if key in self.check_in_conos["rows_individual"]:
                    ws.cell(row=2, column=col).fill = (
                        self.styles["green_fill"]
                        if self.check_in_conos["rows_individual"][key]
                        else self.styles["orange_fill"]
                    )

    def run_check_control(self):
        """
        В исходнике функции проверок вызывались снаружи (run_check_control()).
        Здесь предполагаем, что self.check_in_conos/self.control_results уже заполнены
        логикой верхнего уровня. Ничего не меняем, просто оставляем точку расширения.
        """
        try:
            # если извне передали callable – позволяем выполнить
            if callable(getattr(self, "_external_run_check_control", None)):
                getattr(self, "_external_run_check_control")(
                    self.extract_conos, self.check_in_conos, self.control_results
                )
        except Exception as e:
            logging.debug(f"run_check_control skipped: {e}")

    # ------------------------- Phytosanitary Data -------------------------

    def create_phytosanitary_sheet(self, wb):
        """Лист 'Phytosanitary Data': вывод блоков страниц 1 (верт.) и 2 (гор.)."""
        if "Phytosanitary Data" in wb.sheetnames:
            ws_phyt = wb["Phytosanitary Data"]
            ws_phyt.delete_rows(1, ws_phyt.max_row)
        else:
            ws_phyt = wb.create_sheet("Phytosanitary Data")

        page1_data = {}
        page2_data = {}
        for key, data in self.data_phito.items():
            if not isinstance(key, tuple) or len(key) != 2:
                continue
            page, num = key
            if page == "page1_vertical" and isinstance(num, int):
                page1_data[num] = data
            elif page == "page2_horizontal" and isinstance(num, int):
                page2_data[num] = data

        ws_phyt.append(["Page 1 (Vertical)"])
        for num in sorted(page1_data.keys()):
            data = page1_data[num]
            ws_phyt.append([data["name"], data["value"]])

        ws_phyt.append([])
        ws_phyt.append(["Page 2 (Horizontal)"])
        for num in sorted(page2_data.keys()):
            data = page2_data[num]
            ws_phyt.append([data["name"], data["value"]])

        self.auto_adjust_columns(ws_phyt)
        return ws_phyt

    # ------------------------- Phito_Check -------------------------

    def create_system_info_sheet(self, wb):
        """Лист 'Phito_Check': свод проверок фитосанитарного сертификата."""
        if "Phito_Check" in wb.sheetnames:
            ws_sys = wb["Phito_Check"]
            ws_sys.delete_rows(1, ws_sys.max_row)
        else:
            ws_sys = wb.create_sheet("Phito_Check")

        green_fill = PatternFill(start_color="00C800", end_color="00C800", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # Заголовок
        headers = ["Категория проверки", "Проверка", "Статус", "Подробности"]
        ws_sys.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws_sys.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        # Проверка совпадения номеров сертификатов
        self.check_in_phito["check_number"] = (
            self.extract_phito["page1_number"] is not None
            and self.extract_phito["page2_number"] is not None
            and self.extract_phito["page1_number"] == self.extract_phito["page2_number"]
        )

        def add_check_row(category, check_name, check_value, details=""):
            status = "OK" if check_value else "Ошибка"
            fill = green_fill if check_value else orange_fill

            ws_sys.append([category, check_name.replace("_", " ").capitalize(), status, details])
            current_row = ws_sys.max_row
            ws_sys.cell(row=current_row, column=3).fill = fill
            for c in range(1, 5):
                ws_sys.cell(row=current_row, column=c).alignment = wrap_alignment

        # Совпадение номеров
        add_check_row(
            "Основные проверки",
            "Совпадение номеров страниц",
            self.check_in_phito["check_number"],
            f"Номера: '{self.extract_phito['page1_number']}' vs '{self.extract_phito['page2_number']}'",
        )

        # Остальные проверки (наличие ключевых слов, веса, фрукты, даты, N/A-блоки и т.д.)
        main_checks = [
            (
                "check_maritime",
                "Наличие MARITIME",
                lambda: self.data_phito.get(("page1_vertical", 4), {}).get("value", "").upper(),
            ),
            (
                "check_ex_leningrad",
                "Наличие EX LENINGRAD",
                lambda: self.data_phito.get(("page1_vertical", 5), {}).get("value", "").upper(),
            ),
            (
                "check_netto",
                "Совпадение Net Weight",
                f"{self.extract_phito.get('EX_ph1_8_NETTO')} vs {self.extract_phito.get('EX_ph2_6_NETTO')}",
            ),
            (
                "check_gross",
                "Совпадение Gross Weight",
                f"{self.extract_phito.get('EX_ph1_8_GROSS')} vs {self.extract_phito.get('EX_ph2_7_GROSS')}",
            ),
            (
                "check_fruit_es",
                "Совпадение фруктов (исп.)",
                f"'{self.extract_phito.get('EX_ph1_8_fruta')}' vs '{self.extract_phito.get('EX_ph2_4_fruta')}'",
            ),
            (
                "check_fruit_la",
                "Совпадение фруктов (лат.)",
                f"'{self.extract_phito.get('EX_ph1_9_fructus')}' vs '{self.extract_phito.get('EX_ph2_4_fructus')}'",
            ),
            (
                "check_fresh",
                "Наличие FRESH",
                f"'{self.extract_phito.get('EX_ph1__15_fresh')}' vs '{self.extract_phito.get('EX_ph2__15_fresh')}'",
            ),
            (
                "rows_equal",
                "Совпадение количества строк",
                f"{self.rows_count['phito'].get('total_cont_PH1')} vs {self.rows_count['phito'].get('total_cont_PH2')}",
            ),
            (
                "containers_match",
                "Совпадение контейнеров",
                f"{len(self.extract_phito.get('Containers_PH_1', []))} vs {len(self.extract_phito.get('Containers_PH_2', []))}",
            ),
            (
                "numbers_match",
                "Совпадение номеров сертификатов",
                f"'{self.extract_phito.get('page1_number')}' vs '{self.extract_phito.get('page2_number')}'",
            ),
            (
                "boxes_match",
                "Совпадение Box",
                f"{self.extract_phito.get('PH1_Box')} vs {self.extract_phito.get('PH2_Box')}",
            ),
            (
                "BOXES_match",
                "Совпадение BOXES",
                f"{self.extract_phito.get('PH1_BOXES')} vs {self.extract_phito.get('PH2_BOXES')}",
            ),
            ("all_boxes_match", "Полное совпадение Box/BOXES", ""),
            (
                "check_date_PHITO",
                "Совпадение дат фитосанитарки",
                f"'{self.extract_phito.get('EX_1_17_Date_PHITO')}' vs '{self.extract_phito.get('EX_2_17_Date_PHITO')}'",
            ),
        ]

        for check_key, check_name, details_func in main_checks:
            if isinstance(details_func, str):
                details = details_func
            else:
                text = details_func()
                found = ("MARITIME" in text) if check_key == "check_maritime" else ("EX LENINGRAD" in text)
                details = f"Найдено: {'ДА' if found else 'НЕТ'}"

            add_check_row("Основные проверки", check_name, self.check_in_phito.get(check_key, False), details)

        # N/A-блоки — страница 1
        for i in range(10, 15):
            block_name = {
                10: "Date",
                11: "Treatment",
                12: "Chemical",
                13: "Duration and temperature",
                14: "Concentration",
            }.get(i, f"Блок {i}")
            value = self.check_in_phito.get("check_NA_page1", {}).get(i, False)
            add_check_row(
                "Проверки N/A (верт.)",
                f"{block_name} содержит N/A",
                value,
                f"Блок {i}: {self.data_phito.get(('page1_vertical', i), {}).get('value', '')}",
            )

        # N/A-блоки — страница 2
        for i in range(8, 13):
            block_name = {
                8: "Treatment",
                9: "Duration and temperature",
                10: "Concentration",
                11: "Chemical",
                12: "Date",
            }.get(i, f"Блок {i}")
            value = self.check_in_phito.get("check_NA_page2", {}).get(i, False)
            add_check_row(
                "Проверки N/A (гор.)",
                f"{block_name} содержит N/A",
                value,
                f"Блок {i}: {self.data_phito.get(('page2_horizontal', i), {}).get('value', '')}",
            )

        # Обязательные поля (место выдачи / подпись)
        required_checks = [("place_of_issue", "Место выдачи"), ("authorized_officer", "Подпись ответственного")]
        for field, field_name in required_checks:
            for page in ["page1", "page2"]:
                value = self.check_in_phito.get(f"required_fields_{page}", {}).get(field, False)
                page_type = "верт." if page == "page1" else "гор."
                block_num = 16 if field == "place_of_issue" else 18
                key_tuple = (f"{page}_vertical" if page == "page1" else f"{page}_horizontal", block_num)
                add_check_row(
                    "Обязательные поля",
                    f"{field_name} ({page_type})",
                    value,
                    f"Блок {block_num}: {self.data_phito.get(key_tuple, {}).get('value', '')}",
                )

        # Ширины, фиксация, фильтр
        ws_sys.column_dimensions["A"].width = 25
        ws_sys.column_dimensions["B"].width = 35
        ws_sys.column_dimensions["C"].width = 12
        ws_sys.column_dimensions["D"].width = 50
        ws_sys.freeze_panes = "A2"
        ws_sys.auto_filter.ref = f"A1:D{ws_sys.max_row}"
        return ws_sys

    # ------------------------- Finish_Check -------------------------

    def create_finish_check_sheet(self, wb):
        """Лист 'Finish_Check': сравнение коносамента и фитосертификата + контейнеры."""
        if "Finish_Check" in wb.sheetnames:
            ws_finish = wb["Finish_Check"]
            ws_finish.delete_rows(1, ws_finish.max_row)
        else:
            ws_finish = wb.create_sheet("Finish_Check")

        # 1) Заголовки
        self.setup_finish_check_headers(ws_finish)

        # 2) Пары строк сравнения (SHIPPER, CONSIGNEE, VESSEL, PORT, DATE)
        self.add_comparison_data(ws_finish)

        # 3) Заголовок для списка контейнеров
        ws_finish.append(["finish_conos", "containers_conos", "containers_phito", "finish_phito"])

        # 4) Сами контейнеры
        self.add_container_lists(ws_finish)

        # 5) Итоговые сравнения (сводные переменные)
        self.add_final_comparison_data(ws_finish)

        # 6) Форматирование и подсветки
        self.format_finish_check_sheet(ws_finish)

        return ws_finish

    def setup_finish_check_headers(self, ws):
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:D1")
        ws["A1"] = "КОНОСАМЕНТ"
        ws["C1"] = "ФИТКА"
        ws["A1"].alignment = self.styles["center_alignment"]
        ws["C1"].alignment = self.styles["center_alignment"]
        ws["A1"].font = self.styles["bold_font"]
        ws["C1"].font = self.styles["bold_font"]

    def add_comparison_data(self, ws):
        # 2.1 SHIPPER / Exporter
        conos_shipper = self.data_conos["title"].get((1, 3), {})
        phyto_exporter = self.data_phito.get(("page1_vertical", 1), {})
        ws.append(
            [
                conos_shipper.get("name", "SHIPPER (1)"),
                conos_shipper.get("value", ""),
                phyto_exporter.get("value", ""),
                phyto_exporter.get("name", "1. Name and address of exporter"),
            ]
        )

        # 2.2 CONSIGNEE
        conos_consignee = self.data_conos["title"].get((1, 4), {})
        phyto_consignee = self.data_phito.get(("page1_vertical", 2), {})
        ws.append(
            [
                conos_consignee.get("name", "CONSIGNEE (1)"),
                conos_consignee.get("value", ""),
                phyto_consignee.get("value", ""),
                phyto_consignee.get("name", "2. Declared name and address of consignee"),
            ]
        )

        # 2.3 VESSEL AND VOYAGE NO
        conos_vessel = self.data_conos["title"].get((1, 6), {})
        phyto_vessel = self.data_phito.get(("page1_vertical", 6), {})
        ws.append(
            [
                conos_vessel.get("name", "VESSEL AND VOYAGE NO (1)"),
                conos_vessel.get("value", ""),
                phyto_vessel.get("value", ""),
                phyto_vessel.get("name", "6. Distinguishing marks"),
            ]
        )

        # 2.4 PORT DISCHARGE
        conos_port = self.data_conos["title"].get((1, 8), {})
        phyto_port = self.data_phito.get(("page1_vertical", 5), {})
        ws.append(
            [
                conos_port.get("name", "PORT DISCHARGE (1)"),
                conos_port.get("value", ""),
                phyto_port.get("value", ""),
                phyto_port.get("name", "5. Declared point of entry"),
            ]
        )

        # 2.5 PLACE AND DATE OF ISSUE (в коносе — shipped on board date)
        conos_issue = self.data_conos["title"].get((1, 11), {})
        phyto_issue = self.data_phito.get(("page1_vertical", 17), {})
        ws.append(
            [
                conos_issue.get("name", "SHIPPED ON BOARD DATE (1)"),
                conos_issue.get("value", ""),
                phyto_issue.get("value", ""),
                phyto_issue.get("name", "17. Place and date of issue"),
            ]
        )

    def add_container_lists(self, ws):
        containers_conos = self.finish_conos.get("containers_conos", [])
        containers_phito = self.finish_phito.get("containers_phito", [])

        max_containers = max(len(containers_conos), len(containers_phito))
        for i in range(max_containers):
            ws.append(
                [
                    i + 1,
                    containers_conos[i] if i < len(containers_conos) else "",
                    containers_phito[i] if i < len(containers_phito) else "",
                    i + 1,
                ]
            )

    def add_final_comparison_data(self, ws):
        conos_vars = [
            "total_cont_conos",
            "boxes_conos",
            "gross_perbox_conos",
            "gross_conos",
            "fruit_eng_conos",
            "fruit_esp_conos",
            "fruit_la_conos",
            "fresh_conos",
        ]
        phito_vars = [
            "total_cont_phito",
            "boxes_phito",
            "gross_perbox_phito",
            "gross_phito",
            "fruit_eng_phito",
            "fruit_es_phito",
            "fruit_la_phito",
            "fresh_phito",
        ]

        for conos_var, phito_var in zip(conos_vars, phito_vars):
            conos_value = self.finish_conos.get(conos_var, "")
            phito_value = self.finish_phito.get(phito_var, "")

            if isinstance(conos_value, tuple):
                conos_value = ", ".join(conos_value)
            if isinstance(phito_value, tuple):
                phito_value = ", ".join(phito_value)

            ws.append([conos_var, conos_value, phito_value, phito_var])

    def format_finish_check_sheet(self, ws):
        """Полное форматирование Finish_Check, включая подсветки и комментарии."""
        green_fill = PatternFill(start_color="00C800", end_color="00C800", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        bold_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Заголовки шапки
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:D1")
        for cell in ["A1", "C1"]:
            ws[cell].alignment = center_alignment
            ws[cell].font = bold_font
            ws[cell].fill = header_fill

        # Поиск ключевых строк
        for row in ws.iter_rows():
            # заголовок таблицы контейнеров
            if row[0].value == "finish_conos" and row[1].value == "containers_conos":
                for cell in row:
                    cell.font = bold_font
                    cell.fill = header_fill

            # количество контейнеров
            elif row[0].value == "total_cont_conos":
                val_conos = ws.cell(row=row[0].row, column=2).value
                val_phito = ws.cell(row=row[0].row, column=3).value

                if val_conos is None or val_phito is None:
                    fill_color = red_fill
                elif val_conos == val_phito:
                    fill_color = green_fill
                else:
                    fill_color = orange_fill

                ws.cell(row=row[0].row, column=1).fill = fill_color
                ws.cell(row=row[0].row, column=4).fill = fill_color

                if val_conos != val_phito:
                    msg = (
                        "Количество контейнеров не совпадает"
                        if all(v is not None for v in [val_conos, val_phito])
                        else "Отсутствует значение"
                    )
                    ws.cell(row=row[0].row, column=1).comment = Comment(msg, "Система")

            # количество коробок (boxes)
            elif row[0].value == "boxes_conos":
                val_conos = ws.cell(row=row[0].row, column=2).value
                val_phito = ws.cell(row=row[0].row, column=3).value

                is_valid = self.control_results.get("control_check", {}).get("CC_boxes", False)

                if val_conos is None or val_phito is None:
                    fill_color = red_fill
                elif is_valid:
                    fill_color = green_fill
                else:
                    fill_color = orange_fill

                ws.cell(row=row[0].row, column=1).fill = fill_color
                ws.cell(row=row[0].row, column=4).fill = fill_color

                if not is_valid:
                    ws.cell(
                        row=row[0].row, column=1
                    ).comment = Comment(
                        f"Количество коробок не совпадает: {val_conos} vs {val_phito}", "Система"
                    )

        # Контейнеры (таблица ниже)
        container_count = 0
        for row in ws.iter_rows(min_row=8):
            if isinstance(row[0].value, int):
                container_count += 1
                cont_conos = row[1].value
                cont_phito = row[2].value
                if cont_conos and cont_phito:
                    fill_color = green_fill if cont_conos == cont_phito else orange_fill
                else:
                    fill_color = red_fill

                ws.cell(row=row[0].row, column=1).fill = fill_color
                ws.cell(row=row[0].row, column=4).fill = fill_color

                if cont_conos != cont_phito:
                    msg = (
                        "Контейнеры не совпадают"
                        if cont_conos and cont_phito
                        else "Отсутствует значение контейнера"
                    )
                    ws.cell(row=row[0].row, column=1).comment = Comment(msg, "СистемA")
            else:
                break

        # Блок SHIPPER/CONSIGNEE/VESSEL/PORT/DATE — подсветки по control_results
        row_mapping = {
            2: ("Exporter", "SHIPPER", "CC_Exporter"),
            3: ("Consignee", "CONSIGNEE", "CC_Consignee"),
            4: ("Vessel_Voyage", "VESSEL AND VOYAGE NO", "CC_Vessel_Voyage"),
            5: ("Port_Discharge", "PORT DISCHARGE", "CC_Port_Discharge"),
            6: ("Date", "SHIPPED ON BOARD DATE", "CC_date"),
        }
        for row_idx, (result_key, _label, check_key) in row_mapping.items():
            try:
                data = self.control_results.get(result_key, {})
                control_ok = self.control_results.get("control_check", {}).get(check_key, False)
                color = green_fill if control_ok else orange_fill
                ws.cell(row=row_idx, column=1).fill = color
                ws.cell(row=row_idx, column=4).fill = color

                if data.get("unmatched_words_conos"):
                    cell = ws.cell(row=row_idx, column=2)
                    cell.font = Font(color="FFA500")
                    cell.comment = Comment(
                        "Несовпавшие слова в коносаменте:\n" + ", ".join(data["unmatched_words_conos"]),
                        "Система",
                    )

                if data.get("unmatched_words_phito"):
                    cell = ws.cell(row=row_idx, column=3)
                    cell.font = Font(color="FFA500")
                    cell.comment = Comment(
                        "Несовпавшие слова в фитосанитарке:\n" + ", ".join(data["unmatched_words_phito"]),
                        "Система",
                    )
            except Exception as e:
                logging.error(f"Ошибка форматирования строки {row_idx}: {e}")

        # Прочие своды (gross/fruit/fresh)
        control_mapping = {
            "gross_perbox_conos": "CC_gross_perbox",
            "gross_conos": "CC_gross_weight",
            "fruit_eng_conos": "CC_fruit",
            "fruit_esp_conos": "CC_fruit",
            "fruit_la_conos": "CC_fruit",
            "fresh_conos": "CC_fresh",
        }
        start_results_row = 8 + container_count + 2
        for row in ws.iter_rows(min_row=start_results_row, max_row=ws.max_row):
            key = row[0].value
            if key in control_mapping:
                is_valid = self.control_results.get("control_check", {}).get(control_mapping[key], False)
                val_conos = ws.cell(row=row[0].row, column=2).value
                val_phito = ws.cell(row=row[0].row, column=3).value

                if val_conos is None or val_phito is None:
                    fill_color = red_fill
                elif is_valid:
                    fill_color = green_fill
                else:
                    fill_color = orange_fill

                ws.cell(row=row[0].row, column=1).fill = fill_color
                ws.cell(row=row[0].row, column=4).fill = fill_color

        # Ширины и общее форматирование
        for col, width in {"A": 25, "B": 35, "C": 35, "D": 25}.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
            ws.row_dimensions[row[0].row].height = None

        ws.freeze_panes = "A2"

    # ------------------------- Утилиты -------------------------

    def safe_int(self, value, default=0):
        try:
            return int(value)
        except (ValueError, TypeError):
            return default

    def safe_float(self, value, default=0.0):
        try:
            return float(str(value).replace(",", "."))
        except (ValueError, TypeError):
            return default

    def align_list(self, lst, max_len, default):
        if lst is None:
            lst = []
        return lst + [default] * (max_len - len(lst)) if len(lst) < max_len else lst

    def auto_adjust_columns(self, ws, fixed_columns=None):
        if fixed_columns is None:
            fixed_columns = {}
        for col in ws.columns:
            column_letter = get_column_letter(col[0].column)
            if column_letter in fixed_columns:
                ws.column_dimensions[column_letter].width = fixed_columns[column_letter]
                continue
            max_length = 0
            for cell in col:
                if cell.value:
                    cell_value = str(cell.value)
                    lines = cell_value.split("\n")
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
            adjusted_width = (max_length + 2) * 1.2
            adjusted_width = min(adjusted_width, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
